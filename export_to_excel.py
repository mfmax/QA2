#!/usr/bin/env python3
"""
Экспорт данных из базы данных QA в Excel
"""
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config
from db import Database


def setup_worksheet_styles(ws):
    """Настройка стилей для worksheet"""
    # Стиль заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Стиль границ
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    return header_fill, header_font, thin_border


def export_qa_pairs(db: Database, output_path: Path, include_stats: bool = True):
    """Экспорт пар Q&A в Excel"""
    
    # Получение данных - теперь все в одной таблице
    cursor = db.conn.cursor()
    cursor.execute("""
        SELECT 
            id,
            dialog_id,
            filename,
            call_direction,
            operator_phone,
            client_phone,
            call_date,
            call_time,
            question,
            answer,
            direction,
            question_type,
            keywords,
            quality_score,
            created_at
        FROM qa_pairs
        ORDER BY created_at DESC
    """)
    
    rows = cursor.fetchall()
    
    if not rows:
        print("❌ Нет данных для экспорта")
        return False
    
    # Создание workbook
    wb = openpyxl.Workbook()
    
    # Лист 1: Пары Q&A
    ws_pairs = wb.active
    ws_pairs.title = "QA Пары"
    
    # Заголовки
    headers = [
        "ID", "ID Диалога", "Файл", "Направление звонка", "Телефон оператора", 
        "Телефон клиента", "Дата звонка", "Время звонка",
        "Вопрос", "Ответ", "Направление Q&A", "Тип вопроса", 
        "Ключевые слова", "Оценка качества", "Дата создания"
    ]
    
    header_fill, header_font, thin_border = setup_worksheet_styles(ws_pairs)
    
    # Запись заголовков
    for col_num, header in enumerate(headers, 1):
        cell = ws_pairs.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Запись данных
    for row_num, row_data in enumerate(rows, 2):
        # Парсинг keywords из JSON
        try:
            keywords = json.loads(row_data[12]) if row_data[12] else []
            keywords_str = ", ".join(keywords) if keywords else ""
        except:
            keywords_str = row_data[12] or ""
        
        values = [
            row_data[0],  # ID
            row_data[1][:16] + "..." if row_data[1] and len(row_data[1]) > 16 else row_data[1],  # Dialog ID
            row_data[2],  # Filename
            row_data[3] or "",  # Call direction
            row_data[4] or "",  # Operator phone
            row_data[5] or "",  # Client phone
            row_data[6] or "",  # Call date
            row_data[7] or "",  # Call time
            row_data[8],  # Question
            row_data[9],  # Answer
            row_data[10],  # Direction (Q&A)
            row_data[11] or "",  # Question type
            keywords_str,  # Keywords
            row_data[13] or "",  # Quality score
            row_data[14]  # Created at
        ]
        
        for col_num, value in enumerate(values, 1):
            cell = ws_pairs.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            
            # Выравнивание и перенос для длинных текстов
            if col_num in [9, 10]:  # Вопрос и Ответ
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(vertical='top')
    
    # Автоширина колонок
    column_widths = {
        1: 8,   # ID
        2: 20,  # Dialog ID
        3: 40,  # Filename
        4: 12,  # Call direction
        5: 15,  # Operator phone
        6: 15,  # Client phone
        7: 12,  # Call date
        8: 10,  # Call time
        9: 60,  # Question
        10: 80, # Answer
        11: 20, # Direction Q&A
        12: 25, # Type
        13: 30, # Keywords
        14: 12, # Score
        15: 20  # Date
    }
    
    for col_num, width in column_widths.items():
        ws_pairs.column_dimensions[get_column_letter(col_num)].width = width
    
    # Высота строки заголовков
    ws_pairs.row_dimensions[1].height = 30
    
    # Лист 2: Статистика (если требуется)
    if include_stats:
        ws_stats = wb.create_sheet("Статистика")
        stats = db.get_statistics()
        
        # Заголовок - сначала записываем значение, потом объединяем
        ws_stats['A1'].value = "СТАТИСТИКА ИЗВЛЕЧЕНИЯ Q&A"
        ws_stats['A1'].font = Font(bold=True, size=14)
        ws_stats.merge_cells('A1:B1')
        
        # Данные статистики
        stats_data = [
            ("", ""),
            ("Общие показатели", ""),
            ("Всего обработано файлов", stats['total_files']),
            ("Файлов с бизнес-парами", stats['files_with_pairs']),
            ("Всего извлечено пар", stats['total_pairs']),
            ("Средняя оценка качества", f"{stats['avg_quality_score']}/10"),
            ("", ""),
            ("По направлениям", "Количество"),
        ]
        
        for direction, count in stats.get('by_direction', {}).items():
            stats_data.append((f"  • {direction}", count))
        
        stats_data.append(("", ""))
        stats_data.append(("По типам вопросов", "Количество"))
        
        for q_type, count in stats.get('by_type', {}).items():
            if q_type:
                stats_data.append((f"  • {q_type}", count))
        
        # Запись статистики
        for row_num, (label, value) in enumerate(stats_data, 1):
            ws_stats.cell(row=row_num, column=1).value = label
            
            # Пропускаем запись во вторую колонку первой строки (она объединена)
            if row_num > 1:
                ws_stats.cell(row=row_num, column=2).value = value
            
            # Стили для заголовков разделов
            if label and not label.startswith("  •") and value == "":
                ws_stats.cell(row=row_num, column=1).font = Font(bold=True, size=11)
        
        # Ширина колонок
        ws_stats.column_dimensions['A'].width = 40
        ws_stats.column_dimensions['B'].width = 20
    
    # Сохранение
    wb.save(output_path)
    print(f"✅ Экспорт завершен: {output_path}")
    print(f"📊 Экспортировано записей: {len(rows)}")
    
    return True


def export_by_filename(db: Database, output_path: Path, filename_filter: str):
    """Экспорт пар из конкретного файла"""
    cursor = db.conn.cursor()
    cursor.execute("""
        SELECT 
            question, answer, direction, question_type, keywords, quality_score,
            call_direction, operator_phone, client_phone, call_date, call_time
        FROM qa_pairs
        WHERE filename LIKE ?
        ORDER BY id
    """, (f"%{filename_filter}%",))
    
    rows = cursor.fetchall()
    
    if not rows:
        print(f"❌ Нет данных для файла, содержащего: {filename_filter}")
        return False
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QA Пары"
    
    headers = ["Вопрос", "Ответ", "Направление Q&A", "Тип", "Ключевые слова", "Оценка",
               "Направление звонка", "Телефон оператора", "Телефон клиента", "Дата звонка", "Время звонка"]
    header_fill, header_font, thin_border = setup_worksheet_styles(ws)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    for row_num, row_data in enumerate(rows, 2):
        try:
            keywords = json.loads(row_data[4]) if row_data[4] else []
            keywords_str = ", ".join(keywords)
        except:
            keywords_str = row_data[4] or ""
        
        values = [
            row_data[0],  # Question
            row_data[1],  # Answer
            row_data[2],  # Direction Q&A
            row_data[3] or "",  # Type
            keywords_str,  # Keywords
            row_data[5] or "",  # Score
            row_data[6] or "",  # Call direction
            row_data[7] or "",  # Operator phone
            row_data[8] or "",  # Client phone
            row_data[9] or "",  # Call date
            row_data[10] or ""  # Call time
        ]
        
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 10
    
    wb.save(output_path)
    print(f"✅ Экспорт завершен: {output_path}")
    print(f"📊 Экспортировано записей: {len(rows)}")
    
    return True


def main():
    parser = argparse.ArgumentParser(
        description="Экспорт пар Q&A из базы данных в Excel"
    )
    parser.add_argument(
        '--output', '-o',
        type=str,
        default=f"qa_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        help="Имя выходного файла (по умолчанию: qa_export_YYYYMMDD_HHMMSS.xlsx)"
    )
    parser.add_argument(
        '--no-stats',
        action='store_true',
        help="Не включать лист со статистикой"
    )
    parser.add_argument(
        '--filename',
        type=str,
        help="Фильтр по имени файла (экспорт только из указанного файла)"
    )
    parser.add_argument(
        '--db',
        type=str,
        default=str(config.DB_PATH),
        help=f"Путь к базе данных (по умолчанию: {config.DB_PATH})"
    )
    
    args = parser.parse_args()
    
    # Проверка существования БД
    db_path = Path(args.db)
    if not db_path.exists():
        print(f"❌ База данных не найдена: {db_path}")
        print("Запустите сначала main.py для обработки диалогов")
        return 1
    
    output_path = Path(args.output)
    
    try:
        with Database(db_path) as db:
            if args.filename:
                success = export_by_filename(db, output_path, args.filename)
            else:
                success = export_qa_pairs(db, output_path, not args.no_stats)
            
            return 0 if success else 1
            
    except Exception as e:
        import traceback
        print(f"❌ Ошибка при экспорте: {e}")
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())